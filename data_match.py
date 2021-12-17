#生成最后实验所需要的数据

####根据id对医生的features和density，跟问答数据进行匹配，可以得到病例链接的表，然后在跟医生主页上的信息进行合并####

import numpy as np
import pandas as pd
from pandas import Series,DataFrame
from openpyxl import load_workbook

# path_key = 'F:/haodf-prediction/2019/2019similarity/2019_similarity_76050.xls'
# path_value = 'F:/haodf-prediction/data merge/2019match/2019_match.xls'

# path_key = 'F:/haodf-prediction/data merge/2019match/2019_match.xls'
# path_value = 'F:/haodf-prediction/data merge/2019match/2019url_60000好大夫医生主页信息.xls'
##19年根据case匹配

path_key = 'E:/haodf-prediction/data merge/2020match/2020_match.xls'
path_value = 'E:/haodf-prediction/data merge/2020match/2020_all_id.xls'

key = pd.read_excel(path_key)
value = pd.read_excel(path_value)

# key = load_workbook('F:/haodf-prediction/data merge/2020_all_id.xlsx')
# value = load_workbook('F:/haodf-prediction/data merge/2020_match_processed.xlsx')


# all = pd.merge(value, key, on='case', how='left', sort=False)

all = pd.merge(key, value, on='id', how='right', sort=False)
#on为共同的列名
#how为左连接，左边取全部，右边取部分，没有值则用NaN填充
#sort=False为不需要排序
# index = ['case', '3疾病', '4病情描述', '5que', '6ans', 'num_characters', 'num_punctuations',
#          'num_words', 'num_sentences', 'num_nouns', 'num_verbs', 'num_adjs', 'num_dvs', 'num_tagging1', 'num_tagging2',
#          'num_negative', 'num_positive', 'sentiment_score', 'words_len',  'density', 'DP_similarity', '3name', '4clinical',
#          '5academic', '6hospital', '7department', '8annual_haodf', '9hot', '10satisfaction', '11consulation', '12total_visits',
#          '13vote', '14article', '15star', '16after_diagnosis', '17follow-up', '18gifts', '19Satisfaction_efficacy', '20Attitude_satisfaction']
#2019年匹配用case

index = ['id', '3疾病', '4病情描述', '5que', '6ans', 'num_characters', 'num_punctuations',
         'num_words', 'num_sentences', 'num_nouns', 'num_verbs', 'num_adjs', 'num_dvs', 'num_tagging1', 'num_tagging2',
         'num_negative', 'num_positive', 'sentiment_score', 'words_len',  'density', '3name', '4clinical',
         '5academic', '6hospital', '7department', '8annual_haodf', '9hot', '10satisfaction', '11consulation', '12total_visits',
         '13vote', '14article', '15star', '16after_diagnosis', '17follow-up', '18gifts', '19Satisfaction_efficacy', '20Attitude_satisfaction']
#2020年匹配用id,2020年确实相似性那一项，是因为开始没算，后面直接补充到最终表格的
#合并后的数据只取需要的几列，index为列索引
all.loc[:, index]

all.loc[:, index].to_excel('E:/haodf-prediction/data merge/2020match/2020_final_data.xls', encoding='utf_8' )
#现在支持保存成xlsx






